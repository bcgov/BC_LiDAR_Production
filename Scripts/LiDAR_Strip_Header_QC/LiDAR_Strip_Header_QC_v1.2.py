import tkinter as tk
from tkinter import filedialog, messagebox
import os
import subprocess
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import shutil

"""
This program is created to check LiDAR strip header information after they have been fixed using RFQC or run through Stripalign.
See the help information near the end of the script for more information about how it works and what it checks for.

Questions can be emailed to Spencer Floyd at spencer.floyd@gov.bc.ca
"""

# Function to check if the LAStools directory exists
def check_lastools_directory():
    lastools_directory = r"C:\LAStools\bin"
    if not os.path.exists(lastools_directory):
        messagebox.showerror("LASTools Not Found", "'C:\\LAStools\\bin' not found. Please install LASTools before running this program.")
        root.destroy()  # Close the application if LASTools directory is not found

# Move the text files to their own folder
def move_txt_files(source_folder, destination_folder):
    txt_files = [file for file in os.listdir(source_folder) if file.endswith('.txt')]
    for txt_file in txt_files:
        shutil.move(os.path.join(source_folder, txt_file), os.path.join(destination_folder, txt_file))

# Colour the WKT CRS 'Correct' or 'Incorrect' cells accordingly
def color_correct_incorrect_cells(ws, headers):
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            if cell.value == 'Correct':
                cell.fill = PatternFill(start_color="C7EECF", end_color="C7EECF", fill_type="solid")  # Green fill
            elif cell.value == 'Incorrect':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red fill

# Select a proper directory with LiDAR files
def browse_button():
    global directory_path
    directory_path = filedialog.askdirectory()
    if not directory_path:
        return
    
    lidar_files = [file for file in os.listdir(directory_path) if file.endswith(('.laz', '.las'))]
    incorrectly_named_files = [file for file in lidar_files if not re.match(r'^[1-2]\d{4}_\d_[0-9]{3}_20\d\d_\d{4}_C-[A-Za-z]{4}_utm(7|8|9|10|11)\.(las|laz)$', file)]
    
    if incorrectly_named_files:
        messagebox.showerror("Incorrect File Names", f"{len(incorrectly_named_files)} file(s) are incorrectly named. Please correct the names of these files before running this program again.")
        directory_path = None
        directory_label.config(text="")
    else:
        directory_label.config(text="Selected Directory: " + directory_path)

# Run functions
def run_button():
    if not directory_path:
        messagebox.showerror("No Directory Selected", "Please select a directory first.")
        return

    lidar_header_qc_folder = os.path.join(directory_path, "LiDAR Header QC")
    os.makedirs(lidar_header_qc_folder, exist_ok=True)

    lasinfo_command = f'lasinfo -i "{directory_path}\\*.la?" -cpu64 -cores 8 -no_check -quiet -odir "{lidar_header_qc_folder}" -otxt'
    subprocess.run(lasinfo_command, shell=True)

    # Call your script here passing the lidar_header_qc_folder as an argument
    find_key_value_pairs(lidar_header_qc_folder, key_value_pairs)

    # Move .txt files to LASInfo Text Files folder
    lasinfo_text_files_folder = os.path.join(lidar_header_qc_folder, "LASInfo Text Files")
    os.makedirs(lasinfo_text_files_folder, exist_ok=True)
    move_txt_files(lidar_header_qc_folder, lasinfo_text_files_folder)

    messagebox.showinfo("Process Complete", "LiDAR Header QC process is complete.")
    
    # Open LiDAR Header QC folder
    open_folder(lidar_header_qc_folder)

# Function to open a folder in the file explorer
def open_folder(folder_path):
    try:
        subprocess.Popen(f'explorer "{os.path.realpath(folder_path)}"')
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while opening the folder: {str(e)}")

# Find string values based on finding key strings and find out if they are acceptable or not based on LiDAR BC specifications
def find_key_value_pairs(directory, key_value_pairs):
    wb = Workbook()
    ws = wb.active
    
    headers = ['Filename'] + list(key_value_pairs.keys()) + ['WKT OGC COORDINATE SYSTEM']  # Added new column header
    ws.append(headers)  # Header row
    
    # Regular expression pattern
    pattern = r'^[1-2]\d{4}_\d_[0-9]{3}_20\d\d_\d{4}_C-[A-Za-z]{4}_utm(7|8|9|10|11)\.txt$'
    
    # Bold font for headers
    for cell in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for header_cell, header in zip(cell, headers):
            header_cell.value = header
            header_cell.font = Font(bold=True)
    
    for filename in os.listdir(directory):
        if filename.endswith(".txt"):
            filepath = os.path.join(directory, filename)
            row_data = [filename] + [''] * (len(headers) - 1)  # Fill the row_data with empty strings for missing data
            with open(filepath, 'r') as file:
                lines = file.readlines()
                for i, line in enumerate(lines):
                    for key, value in key_value_pairs.items():
                        if key in line:
                            index = line.find(key)
                            found_value = line[index+len(key):].strip().strip("'")  # Remove leading/trailing single quotes
                            row_data[headers.index(key)] = found_value
                    # Additional check for 'LiDAR BC VLR'
                    if 'OP24BMRS001' in open(filepath).read() and 'province_bc' in open(filepath).read():
                        row_data[headers.index('LiDAR BC VLR')] = 'Present'
                    else:
                        row_data[headers.index('LiDAR BC VLR')] = 'Missing'
                    # Extract value for 'WKT OGC COORDINATE SYSTEM'
                    if 'WKT OGC COORDINATE SYSTEM:' in line:
                        found_value = ''
                        for next_line in lines[i+1:]:
                            if next_line.startswith(' ' * 4):  # 4 spaces from the left
                                found_value += next_line.strip()  # Append the line to the found value
                            else:
                                break  # Exit loop if the line doesn't start with four spaces
                        row_data[-1] = found_value.strip('"')  # Assign the found value to the last column in the row_data
                        # Check validity of WKT OGC COORDINATE SYSTEM value based on filename
                        if 'utm' in filename:
                            utm_zone = re.search(r'utm(\d+)', filename).group(1)
                            expected_value = {
                                '7': 'COMPD_CS["NAD83(CSRS) / UTM zone 7N + CGVD2013(CGG2013) height",PROJCS["NAD83(CSRS) / UTM zone 7N",GEOGCS["NAD83(CSRS)",DATUM["NAD83_Canadian_Spatial_Reference_System",SPHEROID["GRS 1980",6378137,298.257222101,AUTHORITY["EPSG","7019"]],AUTHORITY["EPSG","6140"]],PRIMEM["Greenwich",0,AUTHORITY["EPSG","8901"]],UNIT["degree",0.0174532925199433,AUTHORITY["EPSG","9122"]],AUTHORITY["EPSG","4617"]],PROJECTION["Transverse_Mercator"],PARAMETER["latitude_of_origin",0],PARAMETER["central_meridian",-141],PARAMETER["scale_factor",0.9996],PARAMETER["false_easting",500000],PARAMETER["false_northing",0],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["Easting",EAST],AXIS["Northing",NORTH],AUTHORITY["EPSG","3154"]],VERT_CS["CGVD2013(CGG2013) height",VERT_DATUM["Canadian Geodetic Vertical Datum of 2013 (CGG2013)",2005,AUTHORITY["EPSG","1127"]],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["Gravity-related height",UP],AUTHORITY["EPSG","6647"]]]',
                                '8': 'COMPD_CS["NAD83(CSRS) / UTM zone 8N + CGVD2013(CGG2013) height",PROJCS["NAD83(CSRS) / UTM zone 8N",GEOGCS["NAD83(CSRS)",DATUM["NAD83_Canadian_Spatial_Reference_System",SPHEROID["GRS 1980",6378137,298.257222101,AUTHORITY["EPSG","7019"]],AUTHORITY["EPSG","6140"]],PRIMEM["Greenwich",0,AUTHORITY["EPSG","8901"]],UNIT["degree",0.0174532925199433,AUTHORITY["EPSG","9122"]],AUTHORITY["EPSG","4617"]],PROJECTION["Transverse_Mercator"],PARAMETER["latitude_of_origin",0],PARAMETER["central_meridian",-135],PARAMETER["scale_factor",0.9996],PARAMETER["false_easting",500000],PARAMETER["false_northing",0],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["Easting",EAST],AXIS["Northing",NORTH],AUTHORITY["EPSG","3155"]],VERT_CS["CGVD2013(CGG2013) height",VERT_DATUM["Canadian Geodetic Vertical Datum of 2013 (CGG2013)",2005,AUTHORITY["EPSG","1127"]],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["Gravity-related height",UP],AUTHORITY["EPSG","6647"]]]',
                                '9': 'COMPD_CS["NAD83(CSRS) / UTM zone 9N + CGVD2013(CGG2013) height",PROJCS["NAD83(CSRS) / UTM zone 9N",GEOGCS["NAD83(CSRS)",DATUM["NAD83_Canadian_Spatial_Reference_System",SPHEROID["GRS 1980",6378137,298.257222101,AUTHORITY["EPSG","7019"]],AUTHORITY["EPSG","6140"]],PRIMEM["Greenwich",0,AUTHORITY["EPSG","8901"]],UNIT["degree",0.0174532925199433,AUTHORITY["EPSG","9122"]],AUTHORITY["EPSG","4617"]],PROJECTION["Transverse_Mercator"],PARAMETER["latitude_of_origin",0],PARAMETER["central_meridian",-129],PARAMETER["scale_factor",0.9996],PARAMETER["false_easting",500000],PARAMETER["false_northing",0],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["Easting",EAST],AXIS["Northing",NORTH],AUTHORITY["EPSG","3156"]],VERT_CS["CGVD2013(CGG2013) height",VERT_DATUM["Canadian Geodetic Vertical Datum of 2013 (CGG2013)",2005,AUTHORITY["EPSG","1127"]],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["Gravity-related height",UP],AUTHORITY["EPSG","6647"]]]',
                                '10': 'COMPD_CS["NAD83(CSRS) / UTM zone 10N + CGVD2013(CGG2013) height",PROJCS["NAD83(CSRS) / UTM zone 10N",GEOGCS["NAD83(CSRS)",DATUM["NAD83_Canadian_Spatial_Reference_System",SPHEROID["GRS 1980",6378137,298.257222101,AUTHORITY["EPSG","7019"]],AUTHORITY["EPSG","6140"]],PRIMEM["Greenwich",0,AUTHORITY["EPSG","8901"]],UNIT["degree",0.0174532925199433,AUTHORITY["EPSG","9122"]],AUTHORITY["EPSG","4617"]],PROJECTION["Transverse_Mercator"],PARAMETER["latitude_of_origin",0],PARAMETER["central_meridian",-123],PARAMETER["scale_factor",0.9996],PARAMETER["false_easting",500000],PARAMETER["false_northing",0],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["Easting",EAST],AXIS["Northing",NORTH],AUTHORITY["EPSG","3157"]],VERT_CS["CGVD2013(CGG2013) height",VERT_DATUM["Canadian Geodetic Vertical Datum of 2013 (CGG2013)",2005,AUTHORITY["EPSG","1127"]],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["Gravity-related height",UP],AUTHORITY["EPSG","6647"]]]',
                                '11': 'COMPD_CS["NAD83(CSRS) / UTM zone 11N + CGVD2013(CGG2013) height",PROJCS["NAD83(CSRS) / UTM zone 11N",GEOGCS["NAD83(CSRS)",DATUM["NAD83_Canadian_Spatial_Reference_System",SPHEROID["GRS 1980",6378137,298.257222101,AUTHORITY["EPSG","7019"]],AUTHORITY["EPSG","6140"]],PRIMEM["Greenwich",0,AUTHORITY["EPSG","8901"]],UNIT["degree",0.0174532925199433,AUTHORITY["EPSG","9122"]],AUTHORITY["EPSG","4617"]],PROJECTION["Transverse_Mercator"],PARAMETER["latitude_of_origin",0],PARAMETER["central_meridian",-117],PARAMETER["scale_factor",0.9996],PARAMETER["false_easting",500000],PARAMETER["false_northing",0],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["Easting",EAST],AXIS["Northing",NORTH],AUTHORITY["EPSG","2955"]],VERT_CS["CGVD2013(CGG2013) height",VERT_DATUM["Canadian Geodetic Vertical Datum of 2013 (CGG2013)",2005,AUTHORITY["EPSG","1127"]],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["Gravity-related height",UP],AUTHORITY["EPSG","6647"]]]'
                            }.get(utm_zone, None)
                            if found_value == expected_value:
                                row_data[-1] = 'Correct'
                            else:
                                row_data[-1] = 'Incorrect'
            ws.append(row_data)
            # Apply cell coloring after appending the row data
            if re.match(pattern, filename):
                ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="C7EECF", end_color="C7EECF", fill_type="solid") # Green fill
            else:
                ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red fill
            for key, value in key_value_pairs.items():
                if key == 'generating software:':
                    if row_data[headers.index(key)] == 'RiPROCESS':
                        ws.cell(row=ws.max_row, column=headers.index(key) + 1).fill = PatternFill(start_color="C7EECF", end_color="C7EECF", fill_type="solid") # Green fill
                    elif re.match(r'^STRIPALIGN', row_data[headers.index(key)]):
                        ws.cell(row=ws.max_row, column=headers.index(key) + 1).fill = PatternFill(start_color="C7EECF", end_color="C7EECF", fill_type="solid") # Green fill
                    else:
                        ws.cell(row=ws.max_row, column=headers.index(key) + 1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red fill
                elif row_data[headers.index(key)] == value:
                    ws.cell(row=ws.max_row, column=headers.index(key) + 1).fill = PatternFill(start_color="C7EECF", end_color="C7EECF", fill_type="solid") # Green fill
                elif key == 'LiDAR BC VLR':
                    if row_data[headers.index(key)] == 'Present':
                        ws.cell(row=ws.max_row, column=headers.index(key) + 1).fill = PatternFill(start_color="C7EECF", end_color="C7EECF", fill_type="solid") # Green fill
                    else:
                        ws.cell(row=ws.max_row, column=headers.index(key) + 1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red fill
                else:
                    ws.cell(row=ws.max_row, column=headers.index(key) + 1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red fill
    
    # Auto-fit column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column if cell.value is not None]
        if column:
            max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2) * 1.2  # Adjusting width for a bit of padding
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Color Correct/Incorrect cells
    color_correct_incorrect_cells(ws, headers)    

    # Save workbook
    wb.save(os.path.join(directory, 'LiDAR Header QC Summary.xlsx'))

# Key-value pairs to search for
key_value_pairs = {
    'global_encoding:': '17',
    'project ID GUID data 1-4:': '00000000-0000-0000-0000-000000000000',
    'version major.minor:': '1.4',
    'system identifier:': 'Riegl-VQ1560II',
    'point data format:': '6',
    'scale factor x y z:': '0.01 0.01 0.01',
    'LiDAR BC VLR': '',  # Placeholder for the presence of 'OP24BMRS001' and 'province_BC'
    'generating software:': '(STRIPALIGN|RiPROCESS)'
}

def show_help():
    help_text = """
    This program reads a directory of .las, .laz or a mix of both and uses LASTools program LASInfo to create text files that summarize the header information for each LiDAR file.\n
    Specific string values are searched for within each text file to see if the header meets our specifications.\n
    These values should be correct after files have been produced with RFQC, and should be correct even after run through StripAlign.\n
    Below are the header values checked and the correct value for each:

    Filename - [File Source ID]_[Session_Number]_[GPS Day]_[Year]_[System Serial Number]_[Tail Number]_[UTM_Zone].[file type]

    Global Encoding - 17

    Project ID GUID - 00000000-0000-0000-0000-000000000000

    Version Major.Minor - 1.4

    System Identifier - Riegl-VQ1560II

    Point Data Format - 6

    Scale Factor XYZ - 0.01 0.01 0.01

    LiDAR BC VLR - The program looks for the existence of the strings 'OP24BMRS001' and 'province_BC' which will only exist if the VLR has been added properly

    Generating Software - Both 'RiPROCESS' and a string that starts with 'STRIPALIGN' are accepted

    WKT OGC COORDINATE SYSTEM - Checks against correct string based on the UTM zone number in the filename of the text file. Only shows 'Correct' or 'Incorrect' since the actual string is very long.

    These checks are compiled in a .xlsx file and formated by colour (red for incorrect, green for correct)

    For any questions, email Spencer Floyd at spencer.floyd@gov.bc.ca
    """

    messagebox.showinfo("Help", help_text)

# Check for LASTools directory and shut down program if not found
check_lastools_directory()

# Tkinter setup
root = tk.Tk()
root.geometry("600x175")
root.title("LiDAR Flightline Header QC v1.2")

browse_button = tk.Button(root, text="Select Directory", command=browse_button)
browse_button.pack(pady=10)

directory_label = tk.Label(root, text="")
directory_label.pack()

run_button = tk.Button(root, text="Run", command=run_button)
run_button.pack(pady=10)  # Add padding to create space between buttons

help_button = tk.Button(root, text="Help", command=show_help)
help_button.pack(pady=10)

root.mainloop()